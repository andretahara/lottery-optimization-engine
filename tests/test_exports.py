from openpyxl import load_workbook

from lottery_optimizer.disclaimer import DISCLAIMER
from lottery_optimizer.export import (
    build_report,
    charts,
    export_csv,
    export_excel,
    export_report,
)
from lottery_optimizer.export.excel_exporter import SHEETS
from lottery_optimizer.generators import GenerationConstraints, RandomGenerator
from lottery_optimizer.games import registry


def make():
    spec = registry.get("mega-sena")
    p = RandomGenerator().generate(spec, 4, GenerationConstraints(strategy="all_simple"), seed=1)
    return spec, p


def test_csv_columns_and_rows(tmp_path):
    spec, p = make()
    out = export_csv(p, tmp_path / "jogos.csv", game_id=spec.game_id)
    lines = out.read_text(encoding="utf-8").splitlines()
    assert lines[0].startswith("jogo_id,aposta_id,tamanho,dezena_01")
    assert len(lines) == 5  # header + 4 apostas
    assert lines[1].split(",")[0] == "mega-sena"


def test_report_txt_complete(tmp_path):
    spec, p = make()
    out = export_report(p, spec, tmp_path / "relatorio.txt",
                        algorithm="hybrid", seed=42, timestamp="2026-06-27 10:00")
    text = out.read_text(encoding="utf-8")
    assert text.count(DISCLAIMER) >= 1
    assert "Mega-Sena" in text
    assert "Probabilidade teorica" in text
    assert "Seed: 42" in text
    assert "hybrid" in text


def test_excel_has_all_sheets(tmp_path):
    spec, p = make()
    out = export_excel(p, tmp_path / "jogos.xlsx", spec=spec, algorithm="grasp", seed=7,
                       score_history=[0.1, 0.2, 0.3])
    wb = load_workbook(out)
    assert wb.sheetnames == SHEETS
    aviso = wb["AvisoMatematico"]["A2"].value
    assert "mesma probabilidade individual" in aviso
    assert wb["Jogos"].max_row == 5  # header + 4


def test_charts_png(tmp_path):
    spec, p = make()
    f1 = charts.plot_frequency(p, spec, tmp_path / "frequencia_dezenas.png")
    f2 = charts.plot_score_history([0.1, 0.2, 0.25], tmp_path / "score_history.png")
    f3 = charts.plot_overlap_distribution(p, tmp_path / "distribuicao_sobreposicao.png")
    f4 = charts.plot_coverage_by_subset(p, spec, tmp_path / "cobertura_por_subconjunto.png")
    for f in (f1, f2, f3, f4):
        assert f.exists() and f.stat().st_size > 0


def test_build_report_disclaimer_only_args():
    spec, p = make()
    assert DISCLAIMER in build_report(p, spec)
