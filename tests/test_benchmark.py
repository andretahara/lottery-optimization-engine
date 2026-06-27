from openpyxl import load_workbook

from lottery_optimizer.benchmark import _is_valid, run_benchmark
from lottery_optimizer.core.portfolio import Portfolio
from lottery_optimizer.core.ticket import Ticket
from lottery_optimizer.games import registry

ALGOS = ["random", "balanced_random", "greedy_coverage", "local_search", "hybrid"]


def test_benchmark_runs_and_exports(tmp_path):
    spec = registry.get("quina")
    summary = run_benchmark(spec, budget=8, seeds=[1, 2], algorithms=ALGOS,
                            iterations=30, runtime_seconds=None, out_dir=tmp_path)
    assert summary["winner"] in ALGOS
    for r in summary["rows"]:
        assert r["valid"] is True
        for k in ("score", "improvement", "main_unique", "pair_unique", "triple_unique",
                  "quad_unique", "mean_distance", "elapsed", "peak_kb"):
            assert k in r
    for f in ["benchmark.csv", "benchmark.xlsx", "benchmark_report.txt",
              "benchmark_score.png", "benchmark_coverage.png"]:
        assert (tmp_path / f).exists(), f
    assert "mesma probabilidade" in (tmp_path / "benchmark_report.txt").read_text(encoding="utf-8")


def test_winner_is_best_valid_by_score(tmp_path):
    spec = registry.get("quina")
    s = run_benchmark(spec, 8, [1], ALGOS, 30, None, tmp_path)
    valid = [r for r in s["rows"] if r["valid"]]
    best = max(valid, key=lambda r: r["score"])
    assert s["winner"] == best["algorithm"]


def test_excel_has_aviso_sheet(tmp_path):
    spec = registry.get("quina")
    run_benchmark(spec, 6, [1], ["random", "local_search"], 20, None, tmp_path)
    wb = load_workbook(tmp_path / "benchmark.xlsx")
    assert "AvisoMatematico" in wb.sheetnames


def test_validity_excludes_bad_portfolio():
    spec = registry.get("quina")
    good = Portfolio(spec, [Ticket(numbers=(1, 2, 3, 4, 5)), Ticket(numbers=(6, 7, 8, 9, 10))])
    assert _is_valid(good, spec, 2) is True
    assert _is_valid(good, spec, 3) is False        # orcamento != tamanho da carteira
