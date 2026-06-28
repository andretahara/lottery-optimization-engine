"""Verificacao ponta-a-ponta dos 11 itens de prontidao (jogo ARTIFICIAL pequeno).

Cristaliza a prova manual feita em 2026-06-27: geracao -> exports (CSV/Excel/TXT/PNG)
-> sem duplicatas -> orcamento respeitado -> probabilidade = K-subsets unicos / C(N,K)
-> reprodutibilidade por seed. Jogo ficticio, preco marcado example (nunca oficial).
"""
from decimal import Decimal
from fractions import Fraction

import pytest
from openpyxl import load_workbook

from lottery_optimizer.core.combinations import n_choose_k
from lottery_optimizer.core.cost import CostModel
from lottery_optimizer.core.coverage import CombinationCoverage
from lottery_optimizer.core.game import GameSpec
from lottery_optimizer.disclaimer import DISCLAIMER
from lottery_optimizer.export import charts, export_csv, export_excel, export_report
from lottery_optimizer.export.excel_exporter import SHEETS
from lottery_optimizer.generators import GENERATORS, GenerationConstraints

SEED = 12345
BUDGET = 8


@pytest.fixture
def spec():
    return GameSpec(
        game_id="mini-art",
        name="Mini Artificial",
        universe_min=1,
        universe_max=20,
        draw_size=5,
        allowed_ticket_sizes=(5,),
        price_table={5: Decimal("2.50")},
        price_status="example",
        price_source_note="EXAMPLE_NOT_OFFICIAL",
    )


def _gen(spec):
    return GENERATORS["hybrid_initial"]().generate(
        spec, BUDGET, GenerationConstraints(strategy="all_simple"), SEED
    )


def test_item3_generate_small_portfolio(spec):
    p = _gen(spec)
    assert len(p) == BUDGET
    assert all(len(t.numbers) == spec.draw_size for t in p)


def test_item4_csv_export(spec, tmp_path):
    p = _gen(spec)
    path = export_csv(p, tmp_path / "jogos.csv", game_id=spec.game_id)
    lines = path.read_text(encoding="utf-8").splitlines()
    assert lines[0].startswith("jogo_id,aposta_id,tamanho,dezena_01")
    assert len(lines) == BUDGET + 1


def test_item5_excel_export(spec, tmp_path):
    p = _gen(spec)
    path = export_excel(p, tmp_path / "jogos.xlsx", spec=spec, algorithm="hybrid_initial", seed=SEED)
    wb = load_workbook(path)
    assert wb.sheetnames == SHEETS


def test_item6_txt_report_has_disclaimer(spec, tmp_path):
    p = _gen(spec)
    cm = CostModel(spec)
    path = export_report(
        p, spec, tmp_path / "relatorio.txt", cost_model=cm,
        budget=Decimal(BUDGET) * Decimal("2.50"), algorithm="hybrid_initial", seed=SEED,
    )
    assert DISCLAIMER in path.read_text(encoding="utf-8")


def test_item7_charts(spec, tmp_path):
    p = _gen(spec)
    pngs = [
        charts.plot_frequency(p, spec, tmp_path / "freq.png"),
        charts.plot_overlap_distribution(p, tmp_path / "overlap.png"),
        charts.plot_coverage_by_subset(p, spec, tmp_path / "cov.png"),
    ]
    for f in pngs:
        assert f.exists() and f.stat().st_size > 0


def test_item8_no_duplicates(spec):
    p = _gen(spec)
    assert len({t.numbers for t in p}) == len(p)


def test_item9_cost_within_budget(spec):
    p = _gen(spec)
    cm = CostModel(spec)
    budget = Decimal(BUDGET) * Decimal("2.50")
    cost = cm.portfolio_cost(p)
    assert cost.amount <= budget
    assert cost.is_estimate is True  # preco example, nunca oficial


def test_item10_probability_unique_over_cnk(spec):
    p = _gen(spec)
    cov = CombinationCoverage(spec)
    unique = cov.count_unique(p, spec.draw_size, mode="exact")
    total = n_choose_k(spec.pool, spec.draw_size)  # C(20,5)
    expected = Fraction(unique, total)
    assert expected <= 1
    assert unique == BUDGET  # apostas simples distintas, 1 K-subset cada
    assert total == 15504


def test_item11_reproducible_same_seed(spec, tmp_path):
    p1 = _gen(spec)
    p2 = _gen(spec)
    assert [t.numbers for t in p1] == [t.numbers for t in p2]
    a = export_csv(p1, tmp_path / "a.csv", game_id=spec.game_id)
    b = export_csv(p2, tmp_path / "b.csv", game_id=spec.game_id)
    assert a.read_bytes() == b.read_bytes()
