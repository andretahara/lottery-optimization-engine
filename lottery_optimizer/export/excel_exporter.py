"""Export Excel multi-aba (openpyxl): Resumo, Jogos, Frequencias, Metricas, Cobertura,
ScoreHistory, Configuracao, AvisoMatematico. Cabecalho congelado, filtros, largura auto."""

from __future__ import annotations

from pathlib import Path

from ..core.game import GameSpec
from ..core.portfolio import Portfolio
from ..disclaimer import DISCLAIMER
from .report_data import ReportData, build_report_data

SHEETS = ["Resumo", "Jogos", "Frequencias", "Metricas", "Cobertura", "ScoreHistory",
          "Configuracao", "AvisoMatematico"]


def _autofit(ws) -> None:
    from openpyxl.utils import get_column_letter

    for col in ws.columns:
        width = max((len(str(c.value)) for c in col if c.value is not None), default=8)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(width + 2, 60)


def _table(ws, header, rows) -> None:
    ws.append(header)
    for r in rows:
        ws.append(list(r))
    ws.freeze_panes = "A2"
    if rows:
        from openpyxl.utils import get_column_letter
        last = get_column_letter(len(header))
        ws.auto_filter.ref = f"A1:{last}{len(rows) + 1}"
    _autofit(ws)


def export_excel(portfolio: Portfolio, path, *, spec: GameSpec = None, report_data: ReportData = None,
                 **meta) -> Path:
    from openpyxl import Workbook

    if report_data is None:
        if spec is None:
            spec = portfolio.spec
        report_data = build_report_data(portfolio, spec, **meta)
    rd = report_data
    spec = rd.spec
    p = Path(path)
    p.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    wb.remove(wb.active)

    ws = wb.create_sheet("Resumo")
    money = 'R$ #,##0.00'
    _table(ws, ["campo", "valor"], [
        ("Loteria", f"{spec.name} ({spec.game_id})"),
        ("Apostas", rd.n_tickets),
        ("Combinacoes simples equivalentes", rd.equivalent_simple_total),
        ("Cobertura principal unica", rd.unique_main),
        ("Probabilidade premio principal", float(rd.p_main)),
        ("Orcamento", float(rd.budget) if rd.budget is not None else "n/d"),
        ("Custo total", float(rd.total_cost) if rd.total_cost is not None else "n/d"),
        ("Custo e estimativa?", bool(rd.cost_is_estimate)),
        ("Algoritmo", rd.algorithm or "n/d"),
        ("Seed", rd.seed if rd.seed is not None else "n/d"),
    ])
    for row in ws.iter_rows(min_row=2):
        if row[0].value in ("Orcamento", "Custo total") and isinstance(row[1].value, float):
            row[1].number_format = money
        if row[0].value == "Probabilidade premio principal":
            row[1].number_format = "0.000000%"

    ws = wb.create_sheet("Jogos")
    max_size = max((len(t) for t in portfolio), default=0)
    _table(ws, ["aposta_id", "tamanho"] + [f"dezena_{i:02d}" for i in range(1, max_size + 1)],
           [[i, len(t), *t.numbers, *([""] * (max_size - len(t)))] for i, t in enumerate(portfolio, 1)])

    ws = wb.create_sheet("Frequencias")
    _table(ws, ["dezena", "frequencia"], [[n, c] for n, c in sorted(rd.frequency.items())])

    ws = wb.create_sheet("Metricas")
    _table(ws, ["metrica", "valor"], [
        ("entropia", rd.entropy), ("distancia_media", rd.mean_distance),
        ("intersecao_media", rd.mean_intersection), ("intersecao_max", rd.max_intersection),
        ("pares_impares", str(rd.odd_even)), ("baixas_altas", str(rd.low_high)),
    ])

    ws = wb.create_sheet("Cobertura")
    _table(ws, ["subconjunto", "unica", "bruta", "redundancia", "razao_unica_bruta"], [
        ["principal", rd.unique_main, rd.raw_main, rd.raw_main - rd.unique_main,
         rd.unique_main / rd.raw_main if rd.raw_main else 1.0],
        ["pares", rd.pair_cov.unique, rd.pair_cov.raw, rd.pair_cov.redundancy, rd.pair_cov.unique_raw_ratio],
        ["trincas", rd.triple_cov.unique, rd.triple_cov.raw, rd.triple_cov.redundancy, rd.triple_cov.unique_raw_ratio],
        ["quadras", rd.quad_cov.unique, rd.quad_cov.raw, rd.quad_cov.redundancy, rd.quad_cov.unique_raw_ratio],
    ])

    ws = wb.create_sheet("ScoreHistory")
    _table(ws, ["iteracao", "score"], [[i, s] for i, s in enumerate(rd.score_history)])

    ws = wb.create_sheet("Configuracao")
    _table(ws, ["campo", "valor"], [
        ("config_preco", rd.price_config), ("timestamp", rd.timestamp or "n/d"),
        ("modo_cobertura", rd.coverage_mode_used),
        *[(f"param.{k}", str(v)) for k, v in rd.params.items()],
    ])

    ws = wb.create_sheet("AvisoMatematico")
    ws.append(["aviso"])
    ws.append([DISCLAIMER])
    ws.column_dimensions["A"].width = 100

    # garante ordem das abas
    wb._sheets.sort(key=lambda s: SHEETS.index(s.title) if s.title in SHEETS else 99)
    wb.save(p)
    return p
